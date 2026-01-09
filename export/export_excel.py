from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..db.conexion import conectar
import sqlite3

def aplicar_estilo_encabezado(ws):
    """Aplica un diseño profesional a la primera fila"""
    relleno_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fuente_blanca = Font(color="FFFFFF", bold=True)
    alineacion = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = relleno_azul
        cell.font = fuente_blanca
        cell.alignment = alineacion

def exportar_excel(ruta):
    wb = Workbook()
    conn = None
    try:
        conn = conectar()
        cursor = conn.cursor()

        # ------------------ 1. INVENTARIO MATERIA PRIMA ------------------
        ws_mp = wb.active
        ws_mp.title = "Inventario MP"
        ws_mp.append(["Nombre Insumo", "Unidad", "Stock Actual", "Costo Unit.", "Valor Total"])
        cursor.execute("SELECT nombre, unidad, stock_actual, costo_unitario, (stock_actual * costo_unitario) FROM inventario_materia_prima")
        for row in cursor.fetchall(): ws_mp.append(row)
        aplicar_estilo_encabezado(ws_mp)

        # ------------------ 2. CHORIZOS PRODUCIDOS ------------------
        ws_ch = wb.create_sheet("Chorizos Hechos")
        ws_ch.append(["Fecha", "Tipo de Chorizo", "Tanda #", "Kilos", "Unidades"])
        cursor.execute("""
            SELECT t.fecha, r.nombre, t.numero_tanda, t.cantidad_producida, t.unidades
            FROM tandas t JOIN referencias_chorizo r ON r.id = t.referencia_id
            ORDER BY t.fecha DESC
        """)
        for row in cursor.fetchall(): ws_ch.append(row)
        aplicar_estilo_encabezado(ws_ch)

        # ------------------ 3. PROPUESTA DE PRECIOS (COSTEO) ------------------
        # Esta hoja calcula el costo real basado en lo que gastaste en las tandas
        ws_precios = wb.create_sheet("Propuesta de Precios")
        ws_precios.append(["Referencia", "Costo Insumos / Kg", "Margen Sugerido (30%)", "Precio Venta Sugerido"])
        
        cursor.execute("""
            SELECT 
                r.nombre,
                AVG(tm.total / t.cantidad_producida) as costo_por_kg
            FROM tanda_materia_prima tm
            JOIN tandas t ON t.id = tm.tanda_id
            JOIN referencias_chorizo r ON r.id = t.referencia_id
            GROUP BY r.nombre
        """)
        
        for row in cursor.fetchall():
            referencia = row[0]
            costo_kg = row[1] or 0
            margen = costo_kg * 0.30  # Ejemplo 30% de ganancia
            precio_sugerido = costo_kg + margen
            ws_precios.append([referencia, round(costo_kg, 2), round(margen, 2), round(precio_sugerido, 2)])
        
        aplicar_estilo_encabezado(ws_precios)

        # ------------------ 4. HISTORIAL DE MOVIMIENTOS ------------------
        ws_hist = wb.create_sheet("Historial de Movimientos")
        ws_hist.append(["Fecha", "Hora", "Insumo", "Tipo", "Cantidad", "Stock Resultante", "Referencia"])
        cursor.execute("""
            SELECT h.fecha, h.hora, i.nombre, h.tipo_movimiento, h.cantidad, h.stock_resultante, h.referencia
            FROM historial_inventario_materia_prima h
            JOIN inventario_materia_prima i ON i.id = h.materia_prima_id
            ORDER BY h.fecha DESC, h.hora DESC
        """)
        for row in cursor.fetchall(): ws_hist.append(row)
        aplicar_estilo_encabezado(ws_hist)

        # Ajustar ancho de columnas automáticamente
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 2

        wb.save(ruta)
        return True

    except Exception as e:
        print(f"Error detallado: {e}")
        return False
    finally:
        if conn: conn.close()